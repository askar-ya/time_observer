from datetime import datetime
import json
import openpyxl
import os

from openpyxl.writer.excel import save_workbook
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build

SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
SPREADSHEET = '1Zzdv9t7gZ7-PMfudc4kZiMnoQthavRXstbzfJ2t62Gg'


def connection_google():
    """Подключение к google sheets"""
    creds = None
    if os.path.exists("token.json"):
        creds = Credentials.from_authorized_user_file("token.json", SCOPES)
        # If there are no (valid) credentials available, let the user log in.
        if creds.valid is False:
            os.remove('token.json')

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                "credentials.json", SCOPES
            )
            creds = flow.run_local_server(port=0)
        with open("token.json", "w") as token:
            token.write(creds.to_json())

    service = build('sheets', 'v4', credentials=creds)
    sheets = service.spreadsheets()
    return sheets


def read_data(file: str):
    with open(file, 'r', encoding='utf-8') as f:
        return json.load(f)


def wright_data(file: str, data: dict) -> None:
    with open(file, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False)


def load_project(file: str, dep) -> None:
    wb = openpyxl.load_workbook(file)
    sheet = wb.active

    project_name = sheet['A4'].value
    knots_name = []
    knots_num = []

    for n, i in enumerate(sheet['A:A'][5:]):
        i = i.value
        if i is not None:
            if len(i.split('.')) == 2:
                knots_num.append(n)

    for i in knots_num:
        knots_name.append(sheet['B:B'][5:][i].value + " " + sheet['C:C'][5:][i].value)

    deps = get_all_departments()
    dep = deps[dep]
    data = read_data('projects_list.json')
    data[dep][project_name] = knots_name
    wright_data('projects_list.json', data)

    data = read_data('data.json')
    data[project_name] = {}
    for i in knots_name:
        data[project_name][i] = []
    wright_data('data.json', data)


def get_time(user: str):
    times = read_data('users.json')[user]['time']
    count = len(times)
    if count % 2 != 0:
        times.append(datetime.now().timestamp())
    chunks = [times[i:i + 2] for i in range(0, len(times), 2)]
    seconds = 0
    for point in chunks:
        seconds += (point[1] - point[0])

    seconds = int(seconds)

    """weeks, seconds = divmod(seconds, 7 * 24 * 60 * 60)
    hours_all = seconds / 60 / 60
    days, seconds = divmod(seconds, 24 * 60 * 60)"""
    hours, seconds = divmod(seconds, 60 * 60)
    minutes, seconds = divmod(seconds, 60)

    return f'{hours}ч{minutes}м'


def form_data() -> list:
    projects = read_data('data.json')
    projects_name = list(projects)
    values = [['Дата записи', 'отдел', 'Проект', 'Узел', 'Сотрудник', 'Затраченное время (ч)']]

    for project in projects_name:
        knots_name = list(projects[project])

        for knot in knots_name:
            names = []
            times = []
            date = []
            deps = []
            users = projects[project][knot]
            for user in users:
                names.append(user[0])
                times.append(user[1])
                date.append(user[2])
                deps.append(user[3])
            for n, name in enumerate(names):
                values.append([date[n].replace('-', '.'), deps[n], project, knot, name, times[n]])

    return values


def load_on_google():
    sheets = connection_google()

    data = form_data()

    clear = []

    for i in range(6):
        for a in range(300):
            clear.append([])
            clear[a].append('  ')

    sheets.values().update(spreadsheetId=SPREADSHEET,
                           range='Лист1!A1',
                           valueInputOption='USER_ENTERED',
                           body={'values': clear}).execute()

    sheets.values().update(spreadsheetId=SPREADSHEET,
                           range='Лист1!A1',
                           valueInputOption='USER_ENTERED',
                           body={'values': data}).execute()


def create_file():

    data = form_data()
    file_name = 'time.xlsx'

    try:
        wb = openpyxl.load_workbook(file_name)
    except Exception as e:
        print(e)
        wb = openpyxl.Workbook()

        # Удаление листа, создаваемого по умолчанию, при создании документа
        for sheet_name in wb.sheetnames:
            sheet = wb.get_sheet_by_name(sheet_name)
            wb.remove_sheet(sheet)

    ws = wb.create_sheet('Time')

    for n, coll in enumerate(data, 1):
        for i, row in enumerate(coll, 1):
            ws.cell(row=n, column=i).value = row

    save_workbook(wb, file_name)


def get_all_departments() -> list:
    return [
        'Отдел механической сборки',
        'цех сборки КИПиА',
        'отдел монтажа оборудования',
        'отдел технического контроля'
    ]


def search_project_user(q: str, dep_q: int) -> list:
    data = read_data('projects_list.json')
    res = []
    for dep_i, dep in enumerate(data):
        if dep_i == dep_q:
            for project_i, project in enumerate(data[dep]):
                if q in project:
                    res.append({'name': project, 'dep': dep_i, 'project': project_i})

    return res


